import os
import sqlite3
from io import BytesIO
from datetime import date, datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, g, abort
)
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "college.db")

app = Flask(__name__)
app.secret_key = "change-this-secret-key-in-production"

# ----------------------------------------------------------------------
# Developer / support contact shown in the app footer and on every slip
# ----------------------------------------------------------------------
DEVELOPER_NAME = "Rashid Zada"
DEVELOPER_TITLE = "Full Stack Developer"
DEVELOPER_WHATSAPP = "923470983567"          # international format, no +
DEVELOPER_WHATSAPP_DISPLAY = "0347-0983567"
COLLEGE_NAME = "SWAT DEGREE COLLEGE OF TECHNOLOGY"

# ----------------------------------------------------------------------
# Database helpers
# ----------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    first_run = not os.path.exists(DB_PATH)
    db = sqlite3.connect(DB_PATH)
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS admin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            subject TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS courses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            duration TEXT,
            fee REAL NOT NULL DEFAULT 0,
            id_card_fee REAL NOT NULL DEFAULT 0,
            dmc_fee REAL NOT NULL DEFAULT 0,
            exam_fee REAL NOT NULL DEFAULT 0,
            fund_fee REAL NOT NULL DEFAULT 0,
            teacher_id INTEGER,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_no TEXT,
            name TEXT NOT NULL,
            father_name TEXT,
            phone TEXT,
            course_id INTEGER,
            teacher_id INTEGER,
            total_fee REAL NOT NULL DEFAULT 0,
            installment_count INTEGER NOT NULL DEFAULT 1,
            admission_date TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE SET NULL,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            installment_no INTEGER NOT NULL,
            tuition_amount REAL NOT NULL DEFAULT 0,
            id_card_fee REAL NOT NULL DEFAULT 0,
            dmc_fee REAL NOT NULL DEFAULT 0,
            exam_fee REAL NOT NULL DEFAULT 0,
            fund_fee REAL NOT NULL DEFAULT 0,
            due_date TEXT,
            paid INTEGER NOT NULL DEFAULT 0,
            paid_date TEXT,
            FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE
        );
        """
    )
    db.commit()

    if first_run:
        db.execute(
            "INSERT INTO admin (username, password_hash) VALUES (?, ?)",
            ("admin", generate_password_hash("admin123")),
        )
        db.commit()
    db.close()


# ----------------------------------------------------------------------
# Auth
# ----------------------------------------------------------------------
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("admin_id"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db = get_db()
        row = db.execute("SELECT * FROM admin WHERE username = ?", (username,)).fetchone()
        if row and check_password_hash(row["password_hash"], password):
            session["admin_id"] = row["id"]
            session["admin_username"] = row["username"]
            flash("Welcome back!", "success")
            nxt = request.args.get("next") or url_for("dashboard")
            return redirect(nxt)
        flash("Invalid username or password.", "error")
    return render_template("login.html", dev=dev_context())


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def dev_context():
    return {
        "name": DEVELOPER_NAME,
        "title": DEVELOPER_TITLE,
        "whatsapp": DEVELOPER_WHATSAPP,
        "whatsapp_display": DEVELOPER_WHATSAPP_DISPLAY,
        "college": COLLEGE_NAME,
    }


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def format_pk_whatsapp(phone):
    """Normalize a Pakistani phone number into wa.me format (92XXXXXXXXXX)."""
    if not phone:
        return ""
    digits = "".join(ch for ch in phone if ch.isdigit())
    if digits.startswith("0"):
        digits = "92" + digits[1:]
    elif digits.startswith("92"):
        pass
    elif len(digits) == 10:
        digits = "92" + digits
    return digits


def create_payments_for_student(db, student_id, total_fee, installment_count, course, admission_date_str):
    installment_amount = round(total_fee / installment_count, 2)
    try:
        base_date = datetime.strptime(admission_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        base_date = date.today()

    for i in range(1, installment_count + 1):
        due = base_date + timedelta(days=30 * (i - 1)) if i > 1 else base_date + timedelta(days=15)
        id_card = course["id_card_fee"] if course and i == 1 else 0
        dmc = course["dmc_fee"] if course and i == 1 else 0
        exam = course["exam_fee"] if course and i == 1 else 0
        fund = course["fund_fee"] if course and i == 1 else 0
        db.execute(
            """INSERT INTO payments
               (student_id, installment_no, tuition_amount, id_card_fee, dmc_fee, exam_fee, fund_fee, due_date, paid)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)""",
            (student_id, i, installment_amount, id_card, dmc, exam, fund, due.isoformat()),
        )
    db.commit()


# ----------------------------------------------------------------------
# Dashboard
# ----------------------------------------------------------------------
@app.route("/")
@login_required
def dashboard():
    db = get_db()
    total_students = db.execute("SELECT COUNT(*) c FROM students").fetchone()["c"]
    total_teachers = db.execute("SELECT COUNT(*) c FROM teachers").fetchone()["c"]
    total_courses = db.execute("SELECT COUNT(*) c FROM courses").fetchone()["c"]

    collected = db.execute(
        """SELECT COALESCE(SUM(tuition_amount + id_card_fee + dmc_fee + exam_fee + fund_fee), 0) s
           FROM payments WHERE paid = 1"""
    ).fetchone()["s"]
    pending = db.execute(
        """SELECT COALESCE(SUM(tuition_amount + id_card_fee + dmc_fee + exam_fee + fund_fee), 0) s
           FROM payments WHERE paid = 0"""
    ).fetchone()["s"]

    dues = db.execute(
        """SELECT p.*, s.name AS student_name, s.candidate_no, s.phone,
                  c.name AS course_name
           FROM payments p
           JOIN students s ON s.id = p.student_id
           LEFT JOIN courses c ON c.id = s.course_id
           WHERE p.paid = 0
           ORDER BY p.due_date ASC
           LIMIT 25"""
    ).fetchall()

    courses = db.execute(
        """SELECT c.*, t.name AS teacher_name,
                  (SELECT COUNT(*) FROM students st WHERE st.course_id = c.id) AS student_count
           FROM courses c LEFT JOIN teachers t ON t.id = c.teacher_id"""
    ).fetchall()

    today = date.today().isoformat()

    return render_template(
        "dashboard.html",
        total_students=total_students,
        total_teachers=total_teachers,
        total_courses=total_courses,
        collected=collected,
        pending=pending,
        dues=dues,
        courses=courses,
        today=today,
        dev=dev_context(),
    )


# ----------------------------------------------------------------------
# Teachers
# ----------------------------------------------------------------------
@app.route("/teachers", methods=["GET", "POST"])
@login_required
def teachers():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        subject = request.form.get("subject", "").strip()
        if not name:
            flash("Teacher name is required.", "error")
        else:
            db.execute(
                "INSERT INTO teachers (name, phone, subject) VALUES (?, ?, ?)",
                (name, phone, subject),
            )
            db.commit()
            flash("Teacher added.", "success")
        return redirect(url_for("teachers"))

    rows = db.execute(
        """SELECT t.*, (SELECT COUNT(*) FROM courses c WHERE c.teacher_id = t.id) AS course_count
           FROM teachers t ORDER BY t.name"""
    ).fetchall()
    return render_template("teachers.html", teachers=rows, dev=dev_context())


@app.route("/teachers/delete/<int:teacher_id>", methods=["POST"])
@login_required
def delete_teacher(teacher_id):
    db = get_db()
    db.execute("DELETE FROM teachers WHERE id = ?", (teacher_id,))
    db.commit()
    flash("Teacher removed.", "success")
    return redirect(url_for("teachers"))


# ----------------------------------------------------------------------
# Courses
# ----------------------------------------------------------------------
@app.route("/courses", methods=["GET", "POST"])
@login_required
def courses():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        duration = request.form.get("duration", "").strip()
        fee = float(request.form.get("fee") or 0)
        id_card_fee = float(request.form.get("id_card_fee") or 0)
        dmc_fee = float(request.form.get("dmc_fee") or 0)
        exam_fee = float(request.form.get("exam_fee") or 0)
        fund_fee = float(request.form.get("fund_fee") or 0)
        teacher_id = request.form.get("teacher_id") or None
        if not name:
            flash("Course name is required.", "error")
        else:
            db.execute(
                """INSERT INTO courses
                   (name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id),
            )
            db.commit()
            flash("Course added.", "success")
        return redirect(url_for("courses"))

    rows = db.execute(
        """SELECT c.*, t.name AS teacher_name FROM courses c
           LEFT JOIN teachers t ON t.id = c.teacher_id ORDER BY c.name"""
    ).fetchall()
    teacher_rows = db.execute("SELECT * FROM teachers ORDER BY name").fetchall()
    return render_template("courses.html", courses=rows, teachers=teacher_rows, dev=dev_context())


@app.route("/courses/delete/<int:course_id>", methods=["POST"])
@login_required
def delete_course(course_id):
    db = get_db()
    db.execute("DELETE FROM courses WHERE id = ?", (course_id,))
    db.commit()
    flash("Course removed.", "success")
    return redirect(url_for("courses"))


# ----------------------------------------------------------------------
# Students
# ----------------------------------------------------------------------
@app.route("/students", methods=["GET", "POST"])
@login_required
def students():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        father_name = request.form.get("father_name", "").strip()
        phone = request.form.get("phone", "").strip()
        candidate_no = request.form.get("candidate_no", "").strip()
        course_id = request.form.get("course_id") or None
        teacher_id = request.form.get("teacher_id") or None
        total_fee = float(request.form.get("total_fee") or 0)
        installment_count = int(request.form.get("installment_count") or 1)
        admission_date = request.form.get("admission_date") or date.today().isoformat()

        if not name:
            flash("Student name is required.", "error")
            return redirect(url_for("students"))

        cur = db.execute(
            """INSERT INTO students
               (candidate_no, name, father_name, phone, course_id, teacher_id,
                total_fee, installment_count, admission_date)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (candidate_no, name, father_name, phone, course_id, teacher_id,
             total_fee, installment_count, admission_date),
        )
        db.commit()
        student_id = cur.lastrowid

        course = None
        if course_id:
            course = db.execute("SELECT * FROM courses WHERE id = ?", (course_id,)).fetchone()

        create_payments_for_student(db, student_id, total_fee, installment_count, course, admission_date)
        flash("Student enrolled and fee schedule generated.", "success")
        return redirect(url_for("view_student", student_id=student_id))

    rows = db.execute(
        """SELECT s.*, c.name AS course_name, t.name AS teacher_name,
                  (SELECT COUNT(*) FROM payments p WHERE p.student_id = s.id AND p.paid = 0) AS pending_installments
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           ORDER BY s.created_at DESC"""
    ).fetchall()
    course_rows = db.execute("SELECT * FROM courses ORDER BY name").fetchall()
    teacher_rows = db.execute("SELECT * FROM teachers ORDER BY name").fetchall()
    return render_template(
        "students.html", students=rows, courses=course_rows, teachers=teacher_rows,
        today=date.today().isoformat(), dev=dev_context(),
    )


@app.route("/students/<int:student_id>")
@login_required
def view_student(student_id):
    db = get_db()
    student = db.execute(
        """SELECT s.*, c.name AS course_name, c.duration AS course_duration, t.name AS teacher_name
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           WHERE s.id = ?""",
        (student_id,),
    ).fetchone()
    if not student:
        abort(404)
    payments = db.execute(
        "SELECT * FROM payments WHERE student_id = ? ORDER BY installment_no", (student_id,)
    ).fetchall()
    return render_template("student_view.html", student=student, payments=payments, dev=dev_context())


@app.route("/students/delete/<int:student_id>", methods=["POST"])
@login_required
def delete_student(student_id):
    db = get_db()
    db.execute("DELETE FROM students WHERE id = ?", (student_id,))
    db.commit()
    flash("Student removed.", "success")
    return redirect(url_for("students"))


@app.route("/payments/<int:payment_id>/mark_paid", methods=["POST"])
@login_required
def mark_paid(payment_id):
    db = get_db()
    payment = db.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        abort(404)
    db.execute(
        "UPDATE payments SET paid = 1, paid_date = ? WHERE id = ?",
        (date.today().isoformat(), payment_id),
    )
    db.commit()
    flash("Installment marked as paid.", "success")
    return redirect(url_for("voucher", payment_id=payment_id))


# ----------------------------------------------------------------------
# Voucher / Slip
# ----------------------------------------------------------------------
def get_voucher_data(db, payment_id):
    payment = db.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        return None, None
    student = db.execute(
        """SELECT s.*, c.name AS course_name, c.duration AS course_duration, t.name AS teacher_name
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           WHERE s.id = ?""",
        (payment["student_id"],),
    ).fetchone()
    return payment, student


@app.route("/voucher/<int:payment_id>")
@login_required
def voucher(payment_id):
    db = get_db()
    payment, student = get_voucher_data(db, payment_id)
    if not payment or not student:
        abort(404)
    total = (payment["tuition_amount"] + payment["id_card_fee"] + payment["dmc_fee"]
             + payment["exam_fee"] + payment["fund_fee"])
    wa_number = format_pk_whatsapp(student["phone"])
    wa_message = (
        f"Fee Voucher - {COLLEGE_NAME}\n"
        f"Student: {student['name']}\n"
        f"Candidate No: {student['candidate_no'] or '-'}\n"
        f"Installment: {payment['installment_no']} of {student['installment_count']}\n"
        f"Amount: PKR {total:.0f}\n"
        f"Due Date: {payment['due_date']}\n"
        f"Please find the attached fee voucher PDF."
    )
    return render_template(
        "voucher.html",
        payment=payment,
        student=student,
        total=total,
        wa_number=wa_number,
        wa_message=wa_message,
        today=date.today().strftime("%d/%m/%Y"),
        dev=dev_context(),
    )


@app.route("/voucher/<int:payment_id>/pdf")
@login_required
def voucher_pdf(payment_id):
    from pdf_generator import build_voucher_pdf

    db = get_db()
    payment, student = get_voucher_data(db, payment_id)
    if not payment or not student:
        abort(404)

    buffer = build_voucher_pdf(payment, student, dev_context(), COLLEGE_NAME)
    filename = f"Voucher_{student['candidate_no'] or student['id']}_Inst{payment['installment_no']}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


# ----------------------------------------------------------------------
# Settings (change admin password)
# ----------------------------------------------------------------------
@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    db = get_db()
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        admin = db.execute("SELECT * FROM admin WHERE id = ?", (session["admin_id"],)).fetchone()
        if not check_password_hash(admin["password_hash"], current_password):
            flash("Current password is incorrect.", "error")
        elif len(new_password) < 6:
            flash("New password must be at least 6 characters.", "error")
        elif new_password != confirm_password:
            flash("New passwords do not match.", "error")
        else:
            db.execute(
                "UPDATE admin SET password_hash = ? WHERE id = ?",
                (generate_password_hash(new_password), session["admin_id"]),
            )
            db.commit()
            flash("Password updated successfully.", "success")
        return redirect(url_for("settings"))

    return render_template("settings.html", dev=dev_context())


# ----------------------------------------------------------------------
# Excel Export / Import
# ----------------------------------------------------------------------
@app.route("/data")
@login_required
def data_tools():
    return render_template("data_tools.html", dev=dev_context())


@app.route("/export/excel")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    db = get_db()
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

    # Teachers sheet
    ws = wb.active
    ws.title = "Teachers"
    ws.append(["id", "name", "phone", "subject"])
    for t in db.execute("SELECT id, name, phone, subject FROM teachers ORDER BY id").fetchall():
        ws.append([t["id"], t["name"], t["phone"], t["subject"]])
    style_header(ws, 4)

    # Courses sheet
    ws = wb.create_sheet("Courses")
    ws.append(["id", "name", "duration", "fee", "id_card_fee", "dmc_fee", "exam_fee", "fund_fee", "teacher_id"])
    for c in db.execute(
        "SELECT id, name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id FROM courses ORDER BY id"
    ).fetchall():
        ws.append([c["id"], c["name"], c["duration"], c["fee"], c["id_card_fee"],
                   c["dmc_fee"], c["exam_fee"], c["fund_fee"], c["teacher_id"]])
    style_header(ws, 9)

    # Students sheet
    ws = wb.create_sheet("Students")
    ws.append(["id", "candidate_no", "name", "father_name", "phone", "course_id",
               "teacher_id", "total_fee", "installment_count", "admission_date"])
    for s in db.execute(
        """SELECT id, candidate_no, name, father_name, phone, course_id, teacher_id,
                  total_fee, installment_count, admission_date FROM students ORDER BY id"""
    ).fetchall():
        ws.append([s["id"], s["candidate_no"], s["name"], s["father_name"], s["phone"],
                   s["course_id"], s["teacher_id"], s["total_fee"], s["installment_count"],
                   s["admission_date"]])
    style_header(ws, 10)

    # Payments / Dues sheet
    ws = wb.create_sheet("Payments")
    ws.append(["id", "student_id", "student_name", "installment_no", "tuition_amount",
               "id_card_fee", "dmc_fee", "exam_fee", "fund_fee", "due_date", "paid", "paid_date"])
    for p in db.execute(
        """SELECT p.*, s.name AS student_name FROM payments p
           JOIN students s ON s.id = p.student_id ORDER BY p.student_id, p.installment_no"""
    ).fetchall():
        ws.append([p["id"], p["student_id"], p["student_name"], p["installment_no"],
                   p["tuition_amount"], p["id_card_fee"], p["dmc_fee"], p["exam_fee"],
                   p["fund_fee"], p["due_date"], "Yes" if p["paid"] else "No", p["paid_date"]])
    style_header(ws, 12)

    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"college_data_export_{date.today().isoformat()}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/import/excel", methods=["POST"])
@login_required
def import_excel():
    import openpyxl

    file = request.files.get("excel_file")
    if not file or file.filename == "":
        flash("Please choose an .xlsx file to import.", "error")
        return redirect(url_for("data_tools"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Only .xlsx files are supported.", "error")
        return redirect(url_for("data_tools"))

    db = get_db()
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        flash(f"Could not read the Excel file: {e}", "error")
        return redirect(url_for("data_tools"))

    teacher_map = {}   # old_id -> new_id
    course_map = {}     # old_id -> new_id
    counts = {"teachers": 0, "courses": 0, "students": 0}

    # Teachers
    if "Teachers" in wb.sheetnames:
        ws = wb["Teachers"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[1]:
                continue
            old_id, name, phone, subject = (list(row) + [None] * 4)[:4]
            cur = db.execute(
                "INSERT INTO teachers (name, phone, subject) VALUES (?, ?, ?)",
                (name, phone, subject),
            )
            if old_id is not None:
                teacher_map[int(old_id)] = cur.lastrowid
            counts["teachers"] += 1
        db.commit()

    # Courses
    if "Courses" in wb.sheetnames:
        ws = wb["Courses"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[1]:
                continue
            padded = (list(row) + [None] * 9)[:9]
            old_id, name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, old_teacher_id = padded
            new_teacher_id = teacher_map.get(int(old_teacher_id)) if old_teacher_id not in (None, "") else None
            cur = db.execute(
                """INSERT INTO courses (name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, duration, fee or 0, id_card_fee or 0, dmc_fee or 0, exam_fee or 0, fund_fee or 0, new_teacher_id),
            )
            if old_id is not None:
                course_map[int(old_id)] = cur.lastrowid
            counts["courses"] += 1
        db.commit()

    # Students (+ auto generate payment schedule for each)
    if "Students" in wb.sheetnames:
        ws = wb["Students"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[2]:
                continue
            padded = (list(row) + [None] * 10)[:10]
            (old_id, candidate_no, name, father_name, phone, old_course_id,
             old_teacher_id, total_fee, installment_count, admission_date) = padded
            new_course_id = course_map.get(int(old_course_id)) if old_course_id not in (None, "") else None
            new_teacher_id = teacher_map.get(int(old_teacher_id)) if old_teacher_id not in (None, "") else None
            admission_date_str = str(admission_date)[:10] if admission_date else date.today().isoformat()
            cur = db.execute(
                """INSERT INTO students (candidate_no, name, father_name, phone, course_id, teacher_id,
                                          total_fee, installment_count, admission_date)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (candidate_no, name, father_name, phone, new_course_id, new_teacher_id,
                 total_fee or 0, installment_count or 1, admission_date_str),
            )
            new_student_id = cur.lastrowid
            course = None
            if new_course_id:
                course = db.execute("SELECT * FROM courses WHERE id = ?", (new_course_id,)).fetchone()
            create_payments_for_student(
                db, new_student_id, total_fee or 0, int(installment_count or 1), course, admission_date_str
            )
            counts["students"] += 1
        db.commit()

    flash(
        f"Import complete: {counts['teachers']} teachers, {counts['courses']} courses, "
        f"{counts['students']} students added.", "success"
    )
    return redirect(url_for("data_tools"))


if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="127.0.0.1", port=5000)
